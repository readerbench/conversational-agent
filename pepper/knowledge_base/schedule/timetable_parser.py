from openpyxl import load_workbook
import os
import re
import logging
import codecs
from rdflib import Graph, Literal, URIRef, BNode, Namespace
from rdflib.namespace import RDF, RDFS, XSD
from datetime import datetime

# constants
DAYS = ['Luni', 'Marți', 'Miercuri', 'Joi', 'Vineri']

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)


class SheetParser:
    def __init__(self, sheet, series):
        self.sheet = sheet
        self.series = series

        # borders of the cells will be memoized in a dict for efficiency
        self.cell_border_cache = {}

    @staticmethod
    def xlc(cell):
        """ Build the Excel name representation of a cell. """

        return chr(cell[1] + 64) + str(cell[0])

    def cell_borders(self, cell):
        """
        Find the (manually marked) borders of a rectangle containing the given cell.
        This function considers not only the individual border values of each cell, but also situations that may arise
        from merging cells (when the top-left cell will store the border values for the whole merged rectangle).
        """

        if cell in self.cell_border_cache:
            return self.cell_border_cache[cell]

        row, col = cell
        own_border = self.sheet[self.xlc(cell)].border
        left = own_border.left.style \
               or col > 1 and self.cell_borders((row, col - 1))['right'] \
               or None
        top = own_border.top.style \
              or row > 1 and self.cell_borders((row - 1, col))['bottom'] \
              or None

        is_top_left_merged = False
        right = bottom = None
        for merged_range in self.sheet.merged_cells.ranges:
            if merged_range.min_row <= row <= merged_range.max_row \
                    and merged_range.min_col <= col <= merged_range.max_col:
                # The cell is contained in the current merged range
                is_top_left_merged = self.sheet[self.xlc(cell)] == merged_range.start_cell

                if row == merged_range.max_row:
                    bottom = merged_range.start_cell.border.bottom.style
                if col == merged_range.max_col:
                    right = merged_range.start_cell.border.right.style
                break

        bottom = bottom or self.sheet[self.xlc((row + 1, col))].border.top.style
        if not bottom and not is_top_left_merged:
            bottom = own_border.bottom.style

        right = right or self.sheet[self.xlc((row, col + 1))].border.left.style
        if not right and not is_top_left_merged:
            right = own_border.right.style

        borders = {'left': left, 'top': top, 'right': right, 'bottom': bottom}
        self.cell_border_cache[cell] = borders
        return borders

    def border_box(self, cell):
        """ Identify the minimal rectangle that contains the given cell and is bounded by a border. """

        row, col = cell
        while not self.cell_borders((cell[0], col))['right']:
            col += 1
        while not self.cell_borders((row, cell[1]))['bottom']:
            row += 1
        return cell[0], cell[1], row, col

    def identify_row_hours(self, cell):
        """
        Identify the day and the hour corresponding to each row.
        :return dictionary of pairs row: (day, start hour)
        """

        row_hours = {}
        current_hour = 0
        day_idx = 0
        day = DAYS[day_idx]
        while day:
            if not current_hour:
                # This is the first hour of the current day
                try:
                    current_hour = int(self.sheet[self.xlc(cell)].value.strip().split('-')[0])
                except:
                    logger.error('Parsing of hour range cell failed')
                    raise

            row_hours[cell[0]] = day, current_hour
            current_hour += 1

            if self.cell_borders((cell[0], cell[1] - 1))['bottom']:
                # This is the last hour of the current day
                day_idx += 1
                day = DAYS[day_idx] if day_idx < len(DAYS) else None
                # Restart hour counter
                current_hour = 0

            cell = (cell[0] + 1, cell[1])

        logger.debug(' - Row hours: ' + str(row_hours))
        return row_hours

    def identify_student_groups(self, table_corner):
        """ Identify the student groups/subgroups. """

        cell = (table_corner[0], table_corner[1] + 2)
        groups = []
        while self.sheet[self.xlc(cell)].value:
            # Extract the group name and standardize it
            group_name = self.sheet[self.xlc(cell)].value.strip().upper()
            if re.search('OP', group_name):
                # optional group
                optional_nums = re.findall(r'[0-9]+', group_name)
                group_name = 'OPT' + self.series
            group_name = re.sub(r'\s', '', group_name)  # Remove spaces

            first_col = cell[1]
            # Find the range of columns of the current group
            while not self.cell_borders((cell[0], cell[1]))['right'] and cell[1] < 20:
                cell = (cell[0], cell[1] + 1)

            groups.append((group_name, first_col, cell[1]))

            # Go to the next group
            cell = (cell[0], cell[1] + 1)

        logger.debug(' - Student groups: ' + str(groups))
        return groups

    @staticmethod
    def parse_activity_description(act_description):
        act_type_map = {
            'C': 'course',
            'S': 'seminar',
            'L': 'lab',
            'P': 'p',
        }

        # Clean the full description
        act_description = re.sub(r'\s+', ' ', act_description.strip()).upper()

        act_type = re.search(r'\((S|L|P|C|CURS)\)', act_description)
        act_type = act_type.group(1)[0] if act_type else 'C'

        teacher = re.findall(r'(?:PROF\.?|CONF\.?|[SȘ]\.?L\.?) ?([\w.-]+ [\w.-]+)', act_description)
        rooms = re.findall(r'\w+ ?[0-9]{2,}\w*(?: LEU)?', act_description)  # TODO PR100/101

        # TODO cursuri care nu au titlu, ci doar ID
        title = act_id = None
        if teacher:
            # This is a course
            title = re.match(r'(?:\(.*\) )?([\w ]+)', act_description)
            title = title.group(1) if title else None
        else:
            act_id = re.match(r'[\w ]+', act_description)
            act_id = act_id.group(0) if act_id else None

        act_id = act_id or next(filter(lambda s: s not in ['S', 'L', 'P', 'C', 'CURS'],
                                       re.findall(r'\((\w+)\)', act_description)), None)

        activity = {
            'id': act_id.strip() if act_id else None,
            'title': title.strip() if title else None,
            'type': act_type_map.get(act_type, 'unknown'),
            'teacher': teacher,
            'room': [room.replace(' ', '') for room in rooms]
        }
        return activity

    def identify_activities(self, subtable, row_hours, groups):
        """ Identify activities present in the timetable, together with their time ranges and student groups. """

        row_min, col_min, row_max, col_max = subtable
        schedule = []

        def add_to_schedule(description, time, groups):
            schedule.append({
                **self.parse_activity_description(description),
                'time': (*time, 1) if cell_obj.alignment.horizontal in ['left', None] else time,
                'groups': groups,
                'series': self.series
            })

        for cell in [(r, c) for r in range(row_min, row_max + 1) for c in range(col_min, col_max + 1)]:
            cell_obj = self.sheet[self.xlc(cell)]
            borders = self.cell_borders(cell)
            # Search for activities inspecting bordered boxes by their top-left cell
            if borders['left'] and borders['top'] and \
                    (not cell_obj.fill.patternType or cell_obj.fill.fgColor.tint == 0):
                activity_description = cell_obj.value or ''

                box_top, box_left, box_bottom, box_right = self.border_box(cell)

                # Determine the time when the activity is held
                time = row_hours[box_top][0], \
                       row_hours[box_top][1], \
                       row_hours[box_bottom][1] - row_hours[box_top][1] + 1

                # Determine the groups/semigroups that have this activity
                activity_groups = []
                for group in groups:
                    if box_left <= group[1] and group[2] <= box_right:
                        # The whole group has this activity
                        activity_groups.append((group[0], None))
                    elif group[1] <= box_right < group[2]:
                        # Only the first semigroup has this activity
                        activity_groups.append((group[0], 1))
                    elif group[1] < box_left <= group[2]:
                        # Only the second semigroup has this activity
                        activity_groups.append((group[0], 2))

                # Check if there are different activities on the same slot corresponding to the week parities
                even_week_act = ''
                for r in range(box_top + 1, box_bottom + 1):
                    for c in range(box_left, box_right + 1):
                        if self.sheet[self.xlc((r, c))].value:
                            if (cell_obj.alignment.horizontal in ['left', None] or not activity_description) \
                                    and (c > box_left or self.sheet[self.xlc((r, c))].alignment.horizontal == 'right'):
                                # The value is on the bottom-right side of the box => even parity
                                even_week_act += self.sheet[self.xlc((r, c))].value + ' '
                            else:
                                # The value is just an additional information of the current activity
                                activity_description += ' ' + self.sheet[self.xlc((r, c))].value

                if even_week_act:
                    # Check if there are really 2 activities,
                    # or the second description represents the room of the activity
                    if re.search(r'\(.+\)', activity_description):  # activity title contains its type in brackets ()
                        if re.search(r'\(.+\)', even_week_act):
                            # Both values represent course descriptions
                            act_odd, act_even = activity_description, even_week_act
                        else:
                            # even_week_act is not a standalone activity
                            act_odd, act_even = activity_description + ' ' + even_week_act, None
                    else:
                        # activity_description is not a standalone activity
                        act_odd, act_even = None, even_week_act + ' ' + activity_description

                    if act_odd:
                        add_to_schedule(act_odd, (*time, 1), activity_groups)
                    if act_even:
                        add_to_schedule(act_even, (*time, 2), activity_groups)
                elif activity_description:
                    # No value on the even parity is present
                    add_to_schedule(activity_description,
                                    (*time, 1) if cell_obj.alignment.horizontal in ['left', None] else time,
                                    activity_groups)

        logger.debug(' - Identified activities: ' + '\n'.join([str(item) for item in schedule]))
        return schedule

    def parse(self):
        """ Parse the timetable of a single series. """

        logger.info(' - Parsing schedule for series: ' + self.series)

        # Find the top-left corner of the table
        table_corner = None
        for col in range(1, 10):
            for row in range(1, 30):
                if self.sheet[self.xlc((row, col))].value \
                        and self.sheet[self.xlc((row, col))].value.upper() == 'ZIUA' \
                        and self.sheet[self.xlc((row, col + 1))].value \
                        and self.sheet[self.xlc((row, col + 1))].value.upper() == 'ORA':
                    table_corner = (row, col)
                    break
            if table_corner:
                break

        if not table_corner:
            logger.error('Detecting corner of the table failed')
            return None

        # Move to the first hour range
        cell = (table_corner[0] + 1, table_corner[1] + 1)
        while not self.cell_borders(cell)['top']:
            cell = (cell[0] + 1, cell[1])

        # Identify table headers
        row_hours = self.identify_row_hours(cell)
        groups = self.identify_student_groups(table_corner)

        schedule = self.identify_activities((cell[0], cell[1] + 1, max(row_hours.keys()), groups[-1][2]),
                                            row_hours,
                                            groups)
        return schedule


def parse_timetable(file):
    logger.info('Parsing timetable: ' + file)
    wb = load_workbook(file)

    # Extract the series name from the timetable filename
    series_name = os.path.split(file)[1] \
        .split('.')[0] \
        .upper() \
        .replace('ORAR', '') \
        .replace('MASTER', '-')

    schedule = []
    if len(wb.sheetnames) == 1:
        # single sheet corresponding to the series defined in the file name
        schedule = schedule + SheetParser(wb[wb.sheetnames[0]], series_name).parse()
    else:
        # multiple series grouped in the same Excel file
        for sheetname in wb.sheetnames:
            complete_series_name = (series_name + "-" + sheetname).replace(' ', '')
            schedule = schedule + SheetParser(wb[sheetname], complete_series_name).parse()

    return schedule


def export_schedule_rdf(schedule):
    """
    Export the parsed schedule as an RDF schema in Turtle format (.ttl).
    """

    rdf_graph = Graph()

    # Define prefixes
    base = 'http://www.readerbench.com/pepper#'
    rdf_graph.bind('', base)
    _ = Namespace(base)
    rdf_graph.bind('', _)

    for activity in schedule:
        # Create a unique identifier for the activity
        act_id = URIRef(base + re.sub(r'\s', '_', f'{activity["id"] or ""}_{activity["time"][0]}-{activity["time"][1]}'
                                                  f'_{activity["groups"][0][0]}-{activity["groups"][0][1]}'))
        rdf_graph.add((act_id, RDF.type, _.Activity))

        rdf_graph.add((act_id, _.id, Literal(activity["id"])))
        rdf_graph.add((act_id, _.name, Literal(activity["title"])))
        rdf_graph.add((act_id, _.type, Literal(activity["type"])))

        time = BNode()
        rdf_graph.add((time, _.day, Literal(activity["time"][0])))
        rdf_graph.add((time, _.time, Literal(f'{activity["time"][1]}:00:00', datatype=XSD.time)))
        rdf_graph.add((time, _.duration, Literal(f'PT{activity["time"][2]}H', datatype=XSD.duration)))
        rdf_graph.add((act_id, _.timeSlot, time))

        for (group, semigroup) in activity['groups']:
            group_node = BNode()
            rdf_graph.add((group_node, _.group, Literal(group)))
            rdf_graph.add((group_node, _.semigroup, Literal(semigroup or 0)))
            rdf_graph.add((act_id, _.groups, group_node))
        rdf_graph.add((act_id, _.series, Literal(activity['series'])))

        for room in activity['room']:
            room_node = URIRef(base + room)
            rdf_graph.add((act_id, _.room, room_node))

        for teacher in activity['teacher']:
            rdf_graph.add((act_id, _.teacher, Literal(teacher)))

    with codecs.open("../../microworlds/university_guide/data/kb/schedule.ttl", "w", "utf-8") as rdf_file:
        # Add a description of the data
        rdf_file.write('\n'.join([
            '# This file contains the RDF representation of the courses schedule from',
            '# Faculty of Automatic Control and Computers, University Politehnica of Bucharest',
            '',
            f'# NOTE: This file was autogenerated ({datetime.now()})'
        ]))
        rdf_file.write('\n\n')

        rdf_file.write(rdf_graph.serialize(format='turtle').decode("utf-8"))


def parse_schedule_dir():
    all_schedules = []
    with os.scandir('timetables/') as files:
        for timetable_file in files:
            all_schedules += parse_timetable(timetable_file.path)

    export_schedule_rdf(all_schedules)


if __name__ == "__main__":
    parse_schedule_dir()
